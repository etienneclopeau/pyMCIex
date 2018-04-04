
from demo import Ref, Element, Ensemble
from openpyxl import load_workbook
from collections import OrderedDict
import numpy as np
import copy
import logging
logging.basicConfig(format='%(asctime)s %(levelname)s:%(message)s', 
                    level=logging.INFO, 
                    datefmt='%I:%M:%S')

def get_parameters(wb):
    sheet_param = wb['parametres']
    params =dict()
    params['ntirages'] = sheet_param.cell(row = 33, column = 4).value
    logging.debug('ntirages = %s'%params['ntirages'])
    params['ntirageToDelForProb99.9'] = sheet_param.cell(row = 34, column = 6).value
    logging.debug('ntirageToDelForProb99.9 = %s'%params['ntirageToDelForProb99.9'])

    params['contributeursMethod'] = sheet_param.cell(row = 47, column = 4).value

    return params




def get_ensembles(wb):
    sheet_elements = wb['éléments']

    c_m = 5
    c_m_inerte = c_m + 4
    c_cx = c_m_inerte + 5
    c_cy = c_cx + 4
    c_cz = c_cy + 4
    c_Ix = c_cz + 5
    c_Iy = c_Ix + 4
    c_Iz = c_Iy + 4
    c_Ixy = c_Iz + 5
    c_Ixz = c_Ixy + 4
    c_Iyz = c_Ixz + 4

    print('chargement des elements')
    elements = OrderedDict()
    for row in sheet_elements.iter_rows(row_offset = 4):
        if row[1].value is None: continue
        row = [c.value if c.value is not None else 0 for c in row]
        # row = [c.value for c in row]
        
        
        m = float(row[c_m])
        c = np.array([row[c_cx],row[c_cy],row[c_cz]])/1000.
        I = np.array([[row[c_Ix]  , -row[c_Ixy], -row[c_Ixz] ],
                      [-row[c_Ixy],  row[c_Iy] , -row[c_Iyz] ],
                      [-row[c_Ixz], -row[c_Iyz],  row[c_Iz]  ]])
        garde = dict(m = row[c_m+1],
                     c = np.array([row[c_cx+1],row[c_cy+1],row[c_cz+1]])/1000.,
                     I = np.array([[row[c_Ix+1]  , row[c_Ixy+1], row[c_Ixz+1] ],
                                   [row[c_Ixy+1] ,  row[c_Iy+1], row[c_Iyz+1] ],
                                   [row[c_Ixz+1] , row[c_Iyz+1], row[c_Iz+1]  ]]),
                     )
        def delIfNone(incert):
            if incert['m'] == 0 : incert.pop('m')
            if (incert['c']==0).all() : incert.pop('c')
            if (incert['I']==0).all() : incert.pop('I')
            return incert
        garde = delIfNone(garde)
        meco = dict(m = row[c_m+2],
                     c = np.array([row[c_cx+2],row[c_cy+2],row[c_cz+2]])/1000.,
                     I = np.array([[row[c_Ix+2]  , row[c_Ixy+2], row[c_Ixz+2] ],
                                     [ row[c_Ixy+2],  row[c_Iy+2] , row[c_Iyz+2] ],
                                     [ row[c_Ixz+2],  row[c_Iyz+2],  row[c_Iz+2]  ]]),
                     )
        meco = delIfNone(meco)

        disp = dict(m = row[c_m+3],
                     c = np.array([row[c_cx+3],row[c_cy+3],row[c_cz+3]])/1000.,
                     I = np.array([[row[c_Ix+3]  ,  row[c_Ixy+3],  row[c_Ixz+3] ],
                                     [ row[c_Ixy+3],  row[c_Iy+3] ,  row[c_Iyz+3] ],
                                     [ row[c_Ixz+3],  row[c_Iyz+3],  row[c_Iz+3]  ]]),
                     )
        disp = delIfNone(disp)

        logging.debug("%s : %s, %s, %s,%s, %s, %s"%(row[1],m,c,I,garde,meco,disp))
        # raise
        if row[1] in elements.keys():
            raise(ValueError("l'element %s a déjà été renseigné"%row[1]))
        else: 
            print(row[1])
            if row[1] != 0 : 
                elements[row[1]] = Element(name = row[1] , m = m, c = c, I = I, garde = garde, meco = meco , disp = disp, )


    sheet_referentiel = wb['référentiels']
    print("chargement des référentiels")
    refs = OrderedDict()
    for row in sheet_referentiel.iter_rows(row_offset = 4):
        if row[1].value is None: continue
        row = [c.value if c.value is not None else 0 for c in row]

        if row[1] in refs.keys():
            raise(ValueError("le référentiel %s a déjà été renseigné"%row[1]))
        else:
            print(row[1])
            if row[1] != 0 : 
                if row[8] != 0: ref = refs[row[8]]
                else: ref = None
                refs[row[1]] = Ref(name = row[1], O = row[2:5], psi = row[5], theta = row[6], phi = row[7], ref  = ref,
                           depo = dict( O = row[10:13], psi = row[13], theta = row[14], phi = row[15]))



    sheet_constituants = wb['constituants']
    print("chargement des constituants")
    constituants = OrderedDict()
    for row in sheet_constituants.iter_rows(row_offset = 4):
        if row[1].value is None: continue
        row = [c.value if c.value is not None else 0 for c in row]
        if row[1] in constituants.keys():
            raise(ValueError("le constituant %s a déjà été renseigné"%row[1]))
        else:
            print(row[1])
            if row[1] != 0 : 
                constituants[row[1]] = [elements[row[2]], refs[row[3]]]

    sheet_etats = wb['états']
    print("chargement des états")
    etats = OrderedDict()
    indexCDP = [c.value for c in sheet_etats['A']].index('CDP')
    for col in sheet_etats.iter_cols(min_col = 2):
        col = [c.value if c.value is not None else 0 for c in col]
        if col[2] in etats.keys():
            raise(ValueError("l'état %s a déjà été renseigné"%col[2]))
        else:
            etats[col[2]] = [a  for a in col[3:indexCDP] if a != 0]

    print(etats)

    print("### construction assemblage ###")

    for etat in etats.keys():
        print(  '=====================================')
        print(  '=====================================')

        els = list()
        print(etat)
        for c in etats[etat]:
            print(c)
            # print(constituants[c])
            e = copy.deepcopy(constituants[c][0])
            e.ref = constituants[c][1]
            e.name = c
            els.append(e)
        etats[etat] = Ensemble(*els)

    return etats

# for etat in etats.keys()
#     ensemble = etats[etat]
#     print(ensemble.assemble())
#     # print(ensemble.optimize('Ixx_max'))
#     # raise
#     # print('#######  bornes by optimization #########')
#     # print(ensemble.bornesByOptimization())
#     # raise

#     # print('#######  MTC #########')
#     # m,c,I = ensemble.MTC(ntirages = params['ntirages'])
#     # print('m ',m.shape,m)
#     # print('c ',c.shape, c)
#     # print('I ',I.shape, I)

#     print('########## analitic ###########')
#     print(ensemble.domainAnalitic(approxMethod = True   ))
#     # raise()

def analyseEtat(wb, params, etats):

    sheet_AnalyseEtat = wb['analyse état']
    sae = sheet_AnalyseEtat

    etat = etats[sae['C3'].value]

    def writemcI(m,c,I,frow,fcolumn):
        sae.cell(row = frow, column = fcolumn).value = float(m)
        for i in range(3):
            sae.cell(row = frow, column = fcolumn+1+i).value = c[i]
        for i in range(3):
            sae.cell(row = frow, column = fcolumn+4+i).value = I[i,i]
        sae.cell(row = frow, column = fcolumn+7).value = I[0,1]
        sae.cell(row = frow, column = fcolumn+8).value = I[0,2]
        sae.cell(row = frow, column = fcolumn+9).value = I[1,2]

    etat.nominalAgain(incertAlso = True)
    m,c,I = etat.assemble()
    writemcI(m,c,I,8,5)
    e = etat.domainAnalitic(approxMethod = True   )
    writemcI(e.garde['m'],e.garde['c'],e.garde['I'],9,5)
    writemcI(e.meco['m'],e.meco['c'],e.meco['I'],10,5)
    writemcI(e.disp['m'],e.disp['c'],e.disp['I'],11,5)

    res = etat.bornesByOptimization()
    writemcI(res['mmin'],res['cmin'],res['Imin'],18,5)
    writemcI(res['mmax'],res['cmax'],res['Imax'],19,5)

    #mtc
    m_mtc,c_mtc,I_mtc = etat.MTC(ntirages = params['ntirages'])
    def writeCaraVar(var,i,j):
        sae.cell(row = i, column = j).value = var.mean()
        sae.cell(row = i+1, column = j).value = var.std()
        sae.cell(row = i+2, column = j).value = var.min()
        sae.cell(row = i+3, column = j).value = var.max()
        ordered = np.sort(var)
        sae.cell(row = i+4, column = j).value = ordered[params['ntirageToDelForProb99.9']]
        sae.cell(row = i+5, column = j).value = ordered[-params['ntirageToDelForProb99.9']]
    lfirst,cfirst = 22,5
    for i, var in enumerate([m_mtc,
                            c_mtc[:,0],c_mtc[:,1],c_mtc[:,2],
                            I_mtc[:,0,0],I_mtc[:,1,1],I_mtc[:,2,2],
                            I_mtc[:,0,1],I_mtc[:,0,2],I_mtc[:,1,2]]):
        writeCaraVar(var,lfirst,cfirst+i)
    writeCaraVar(np.sqrt(c_mtc[:,1]**2+c_mtc[:,2]**2),lfirst,cfirst+11)

    
    def max99p9(var):
        ordered = np.sort(var)
        return ordered[-params['ntirageToDelForProb99.9']]

    def getMaxFromMTC():
        m_mtc,c_mtc,I_mtc = etat.MTC(ntirages = params['ntirages'])
        m_maxi = max99p9(m_mtc)
        c_maxi = np.array([max99p9(c_mtc[:,0]),max99p9(c_mtc[:,1]),max99p9(c_mtc[:,2])])
        I_maxi = np.array([max99p9(I_mtc[:,i,j]) for i in range(3) for j in range(3)]).reshape((3,3))
        return m_maxi, c_maxi, I_maxi
    
    def getMax(method = 'MTC'):
        if method == 'MTC':
            return getMaxFromMTC()
        elif method == 'optimize':
            res = etat.bornesByOptimization(only = 'max')
            return res['mmax'],res['cmax'],res['Imax']
        else :
            raise(ValueError)

    # créé la référence
    m_max, c_max, I_max = getMax(method = params['contributeursMethod'])

    # contributeurs
    def contribOfElement(i):
        
        etat.nominalAgain(incertAlso = True)
        
        m_contrib_m = 0 
        try :
            m_contrib_m += etat.elements[i].garde['m']
        except:
            pass
        try:
            m_contrib_m += etat.elements[i].meco['m']
        except:
            pass
        try:
            m_contrib_m +=  etat.elements[i].disp['m']
        except:
            pass


        logging.debug("contrib de la masse de %s"%etat.elements[i].name)
        etat.elements[i].neutraliseIncert('m')
        m_maxi, c_maxi, I_maxi = getMax(method = params['contributeursMethod'])
        c_contrib_m = c_max - c_maxi
        I_contrib_m = I_max - I_maxi


        logging.debug("contrib du centrage de %s"%etat.elements[i].name)
        etat.nominalAgain(incertAlso = True)
        etat.elements[i].neutraliseIncert('c')
        m_maxi, c_maxi, I_maxi = getMax(method = params['contributeursMethod'])
        c_contrib_c = c_max - c_maxi
        I_contrib_c = I_max - I_maxi


        logging.debug("contrib de l'inertie de %s"%etat.elements[i].name)
        etat.nominalAgain(incertAlso = True)
        etat.elements[i].neutraliseIncert('I')
        m_maxi, c_maxi, I_maxi = getMax(method = params['contributeursMethod'])
        I_contrib_I = I_max - I_maxi

        logging.debug("contrib du dépositionnement de %s"%etat.elements[i].name)
        etat.nominalAgain(incertAlso = True)
        # sauvegarde du référentiel d'origine, et remplacement par une copy:
        # cela permet d'avoir la contribution du dépositionnement uniquement par cet élément
        # car plusieurs éléments peuvent utiliser le même référentiel
        # on l'insere aussi dans la liste des rférentiel, juste après le référentiel d'origine
        saveRef = etat.elements[i].ref
        etat.elements[i].ref = copy.deepcopy(saveRef)
        etat.elements[i].ref.name = etat.elements[i].ref.name +'_copy' 
        # etat.findRefs()  
        # indexRef = etat.refs.index(saveRef)+1
        # logging.debug("indexRef %s"%indexRef)
        # logging.debug("len(etat.refs) %s"%len(etat.refs))
        # etat.refs.insert(indexRef, etat.elements[i].ref)
        # logging.debug("len(etat.refs) %s"%len(etat.refs))

        etat.elements[i].neutraliseIncert('depo')
        # logging.debug('verif neutral depo %s'%etat.elements[i].ref.depo)
        m_maxi, c_maxi, I_maxi = getMax(method = params['contributeursMethod'])
        c_contrib_depo = c_max - c_maxi
        I_contrib_depo = I_max - I_maxi
        # on replace le référentiel d'origine, potentiellement identique à d'autres élément
        # et on le suprime de la liste des référentiels
        etat.elements[i].ref = saveRef 
        # logging.debug("indexRef %s"%indexRef)
        # logging.debug("len(etat.refs) %s"%len(etat.refs))
        # etat.refs.pop(indexRef)
  

        etat.nominalAgain(incertAlso = True)
        return m_contrib_m, c_contrib_m, c_contrib_c, c_contrib_depo, I_contrib_m, I_contrib_c, I_contrib_I, I_contrib_depo


    def writeContribs(m_contrib_m, c_contrib_m, c_contrib_c, c_contrib_depo, I_contrib_m, I_contrib_c, I_contrib_I, I_contrib_depo,row):
        sae.cell(row = row, column = 4).value = etat.elements[i].name
        sae.cell(row = row, column = 5).value = m_contrib_m
        sae.cell(row = row, column = 6).value = c_contrib_m[0]
        sae.cell(row = row, column = 7).value = c_contrib_c[0]
        sae.cell(row = row, column = 8).value = c_contrib_depo[0]
        sae.cell(row = row, column = 9).value = c_contrib_m[1]
        sae.cell(row = row, column = 10).value = c_contrib_c[1]
        sae.cell(row = row, column = 11).value = c_contrib_depo[1]
        sae.cell(row = row, column = 12).value = c_contrib_m[2]
        sae.cell(row = row, column = 13).value = c_contrib_c[2]        
        sae.cell(row = row, column = 14).value = c_contrib_depo[2]
        sae.cell(row = row, column = 15).value = I_contrib_m[0,0]
        sae.cell(row = row, column = 16).value = I_contrib_c[0,0]
        sae.cell(row = row, column = 17).value = I_contrib_I[0,0]
        sae.cell(row = row, column = 18).value = I_contrib_depo[0,0]
        sae.cell(row = row, column = 19).value = I_contrib_m[1,1]
        sae.cell(row = row, column = 20).value = I_contrib_c[1,1]
        sae.cell(row = row, column = 21).value = I_contrib_I[1,1]
        sae.cell(row = row, column = 22).value = I_contrib_depo[1,1]
        sae.cell(row = row, column = 23).value = I_contrib_m[2,2]
        sae.cell(row = row, column = 24).value = I_contrib_c[2,2]
        sae.cell(row = row, column = 25).value = I_contrib_I[2,2]
        sae.cell(row = row, column = 26).value = I_contrib_depo[2,2]

    for i in range(len(etat.elements)):
        m_contrib_m, c_contrib_m, c_contrib_c, c_contrib_depo, I_contrib_m, I_contrib_c, I_contrib_I, I_contrib_depo = contribOfElement(i)
        writeContribs(m_contrib_m, c_contrib_m, c_contrib_c, c_contrib_depo, I_contrib_m, I_contrib_c, I_contrib_I, I_contrib_depo, 35+i)








wb = load_workbook('assemblage.xlsx')
wb.get_sheet_names()
params = get_parameters(wb)
etats = get_ensembles(wb)
analyseEtat(wb, params, etats)

wb.save('assemblage_out.xlsx')